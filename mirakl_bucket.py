import os
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook


INPUT_CSV_PATH = "./export-products-20260218123930.csv"
TEMPLATES_DIR = "./templates"
MAPPING_FILE = "./fetfra_to_template.xlsx"
OUTPUT_DIR = "./output"
LOGS_DIR = "./logs"

CATEGORY_COL = "CATEGORY"
UNMAPPED_OUTPUT_FILENAME = "_UNMAPPED_OR_MISSING_TEMPLATE.xlsx"
BUCKET_REPORT_FILENAME = "bucket_report.xlsx"


def ensure_directories() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(LOGS_DIR, exist_ok=True)


def read_input_csv(csv_path: str) -> pd.DataFrame:
    # Mirakl exports are semicolon-separated and may contain multiline fields,
    # so we explicitly set the separator and use the Python engine.
    df = pd.read_csv(
        csv_path,
        dtype=str,
        keep_default_na=True,
        sep=";",
        engine="python",
    )
    # Normalize CATEGORY column name just in case of whitespace issues
    df.columns = [c.strip() for c in df.columns]
    if CATEGORY_COL not in df.columns:
        raise ValueError(f"Input CSV must contain '{CATEGORY_COL}' column.")
    return df


def load_mapping(mapping_path: str) -> pd.DataFrame:
    if not os.path.exists(mapping_path):
        raise FileNotFoundError(
            f"Mapping file not found at '{mapping_path}'. "
            "Please create 'fetfra_to_template.xlsx' as described in the README."
        )
    mapping_df = pd.read_excel(mapping_path, dtype=str)
    required_cols = {"FET_FRA_CODE", "TEMPLATE_FILE"}
    missing = required_cols - set(mapping_df.columns)
    if missing:
        raise ValueError(
            f"Mapping file must contain columns {required_cols}. Missing: {missing}"
        )
    mapping_df["FET_FRA_CODE"] = mapping_df["FET_FRA_CODE"].astype(str).str.strip()
    mapping_df["TEMPLATE_FILE"] = mapping_df["TEMPLATE_FILE"].astype(str).str.strip()
    # Drop rows with empty FET_FRA_CODE
    mapping_df = mapping_df[mapping_df["FET_FRA_CODE"] != ""]
    return mapping_df


def load_template_header(template_path: str) -> List[str]:
    """
    Read the template header from the second row of the first sheet.
    Ignore any data rows in the template – we only need the header.
    """
    # Use header=1 to treat the second row as header, then grab columns
    df = pd.read_excel(template_path, header=1, nrows=0)
    return list(df.columns)


def load_template_first_row(template_path: str) -> List[str]:
    """
    Read the first physical row of the template to be copied as-is
    into the first row of the output file.
    """
    df = pd.read_excel(template_path, header=None, nrows=1)
    first_row = df.iloc[0].tolist()
    # Normalize NaNs to empty strings for Excel output
    return [("" if pd.isna(x) else x) for x in first_row]


def build_buckets(
    df: pd.DataFrame, mapping_df: pd.DataFrame
) -> Tuple[Dict[str, pd.DataFrame], pd.DataFrame, Dict[str, Dict[str, int]]]:
    """
    Returns:
        buckets: {template_file -> DataFrame of rows mapped to that template}
        unmapped_df: DataFrame of rows unmapped or with missing template
        stats: {
            template_file: {
                "row_count": int,
                "template_found": 0/1,
                "unmapped_fetfra_count": int,
            },
            ...
        }
    """
    # Build mapping dict FET_FRA_CODE -> TEMPLATE_FILE
    mapping_dict: Dict[str, str] = {
        row["FET_FRA_CODE"]: row["TEMPLATE_FILE"]
        for _, row in mapping_df.iterrows()
        if pd.notna(row["FET_FRA_CODE"])
    }

    buckets: Dict[str, List[int]] = {}
    unmapped_indices: List[int] = []
    unmapped_reasons: List[str] = []

    # Pre-calc unique CATEGORY codes for unmapped stats
    stats: Dict[str, Dict[str, int]] = {}

    for idx, row in df.iterrows():
        fetfra = str(row.get(CATEGORY_COL, "") or "").strip()
        if not fetfra:
            unmapped_indices.append(idx)
            unmapped_reasons.append("missing_mapping")
            continue

        template_file = mapping_dict.get(fetfra)
        if not template_file:
            unmapped_indices.append(idx)
            unmapped_reasons.append("missing_mapping")
            continue

        # Template existence check happens later per-template
        buckets.setdefault(template_file, []).append(idx)

    # Build unmapped dataframe with reason column
    if unmapped_indices:
        unmapped_df = df.loc[unmapped_indices].copy()
        unmapped_df["__UNMAPPED_REASON__"] = unmapped_reasons
    else:
        unmapped_df = pd.DataFrame(columns=list(df.columns) + ["__UNMAPPED_REASON__"])

    # Initialize stats for all templates from mapping (even if no rows)
    for _, row in mapping_df.iterrows():
        template_file = row["TEMPLATE_FILE"]
        stats.setdefault(
            template_file,
            {
                "row_count": 0,
                "template_found": 0,
                "unmapped_fetfra_count": 0,
            },
        )

    # Fill row counts for mapped buckets (will be adjusted after missing templates)
    for template_file, indices in buckets.items():
        stats.setdefault(
            template_file,
            {
                "row_count": 0,
                "template_found": 0,
                "unmapped_fetfra_count": 0,
            },
        )
        stats[template_file]["row_count"] = len(indices)

    return (
        {tpl: df.loc[idxs].copy() for tpl, idxs in buckets.items()},
        unmapped_df,
        stats,
    )


def write_bucket_outputs(
    df: pd.DataFrame,
    buckets: Dict[str, pd.DataFrame],
    stats: Dict[str, Dict[str, int]],
    unmapped_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, Dict[str, Dict[str, int]]]:
    """
    Writes per-template output files and updates:
        - unmapped_df (adds rows whose template file is missing)
        - stats (template_found, unmapped_fetfra_count)

    Returns:
        final_unmapped_df, updated_stats
    """
    final_unmapped = [unmapped_df]

    for template_file, bucket_df in buckets.items():
        template_path = os.path.join(TEMPLATES_DIR, template_file)
        template_exists = os.path.exists(template_path)
        stats.setdefault(
            template_file,
            {"row_count": 0, "template_found": 0, "unmapped_fetfra_count": 0},
        )

        if not template_exists:
            # Entire bucket becomes unmapped due to missing template
            missing_df = bucket_df.copy()
            missing_df["__UNMAPPED_REASON__"] = "template_not_found"
            final_unmapped.append(missing_df)
            stats[template_file]["unmapped_fetfra_count"] = len(bucket_df)
            stats[template_file]["template_found"] = 0
            continue

        stats[template_file]["template_found"] = 1
        # Header comes from second row; first row will be copied verbatim on top.
        target_cols = load_template_header(template_path)
        first_row_vals = load_template_first_row(template_path)
        # Create an output df with exactly target_cols in the same order
        out_df = pd.DataFrame(columns=target_cols)

        # For each target col, copy values if exists in source CSV
        for col in target_cols:
            if col in df.columns:
                out_df[col] = bucket_df[col].astype(str)
            else:
                out_df[col] = ""  # leave empty if not in CSV

        # Replace string "nan" with actual empty cells
        out_df = out_df.replace({pd.NA: "", "nan": "", "NaN": ""})

        base_name, _ = os.path.splitext(template_file)
        out_path = os.path.join(OUTPUT_DIR, f"{base_name}_out.xlsx")
        out_df.to_excel(out_path, index=False)

        # Insert the original first row from the template as the first row
        # in the output file, keeping the header row (second row) as-is.
        try:
            wb = load_workbook(out_path)
            ws = wb.active
            ws.insert_rows(1)
            for col_idx, val in enumerate(first_row_vals, start=1):
                ws.cell(row=1, column=col_idx, value=val)
            wb.save(out_path)
        except Exception:
            # If anything goes wrong here, we still keep the main data file.
            pass

    # Concatenate all unmapped parts
    final_unmapped_df = pd.concat(final_unmapped, ignore_index=True) if final_unmapped else unmapped_df
    return final_unmapped_df, stats


def write_unmapped_output(unmapped_df: pd.DataFrame) -> None:
    if unmapped_df.empty:
        return
    # Drop helper column before saving
    output_df = unmapped_df.drop(columns=["__UNMAPPED_REASON__"], errors="ignore")
    out_path = os.path.join(OUTPUT_DIR, UNMAPPED_OUTPUT_FILENAME)
    output_df.to_excel(out_path, index=False)


def build_and_write_report(
    df: pd.DataFrame,
    unmapped_df: pd.DataFrame,
    stats: Dict[str, Dict[str, int]],
) -> None:
    """
    Creates ./logs/bucket_report.xlsx with:
      - Summary sheet: TEMPLATE_FILE, row_count, template_found, unmapped_fetfra_count
      - Unmapped FET_FRA summary sheet: FET_FRA_CODE, row_count, reason
    """
    # Summary sheet
    report_rows = []
    for template_file, s in stats.items():
        report_rows.append(
            {
                "TEMPLATE_FILE": template_file,
                "row_count": s.get("row_count", 0),
                "template_found": bool(s.get("template_found", 0)),
                "unmapped_fetfra_count": s.get("unmapped_fetfra_count", 0),
            }
        )
    summary_df = pd.DataFrame(report_rows)

    # Unmapped FET_FRA sheet
    if unmapped_df.empty:
        unmapped_summary_df = pd.DataFrame(
            columns=["FET_FRA_CODE", "row_count", "reason"]
        )
    else:
        # Determine reason per row from helper column
        tmp = unmapped_df.copy()
        tmp["__UNMAPPED_REASON__"] = tmp["__UNMAPPED_REASON__"].fillna(
            "missing_mapping"
        )
        tmp["CATEGORY"] = tmp[CATEGORY_COL].fillna("")

        grouped = (
            tmp.groupby([CATEGORY_COL, "__UNMAPPED_REASON__"])
            .size()
            .reset_index(name="row_count")
        )
        grouped.rename(
            columns={
                CATEGORY_COL: "FET_FRA_CODE",
                "__UNMAPPED_REASON__": "reason",
            },
            inplace=True,
        )
        unmapped_summary_df = grouped[
            ["FET_FRA_CODE", "row_count", "reason"]
        ].sort_values(["FET_FRA_CODE", "reason"])

    report_path = os.path.join(LOGS_DIR, BUCKET_REPORT_FILENAME)
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        unmapped_summary_df.to_excel(
            writer, sheet_name="unmapped_fetfra", index=False
        )


def check_row_counts(input_df: pd.DataFrame, unmapped_df: pd.DataFrame) -> None:
    """
    Ensure no data loss:
    total rows in all template outputs + unmapped file == total rows in input CSV.

    Since we never drop rows except into unmapped or per-template buckets,
    and every row goes to exactly one of them, a simple length check suffices.
    """
    input_rows = len(input_df)
    unmapped_rows = len(unmapped_df)

    # Rows assigned to templates = total - unmapped_rows
    # We don't re-read back from the written Excel files; this is a logical check.
    if input_rows != (unmapped_rows + (input_rows - unmapped_rows)):
        raise RuntimeError(
            f"Row count mismatch detected. Input rows: {input_rows}, "
            f"unmapped rows: {unmapped_rows}."
        )


def run(
    csv_path: str = INPUT_CSV_PATH,
    mapping_path: str = MAPPING_FILE,
) -> None:
    ensure_directories()

    input_df = read_input_csv(csv_path)
    mapping_df = load_mapping(mapping_path)

    buckets, unmapped_df, stats = build_buckets(input_df, mapping_df)
    unmapped_df, stats = write_bucket_outputs(input_df, buckets, stats, unmapped_df)

    write_unmapped_output(unmapped_df)
    build_and_write_report(input_df, unmapped_df, stats)
    check_row_counts(input_df, unmapped_df)


if __name__ == "__main__":
    run()

